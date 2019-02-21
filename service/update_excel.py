import openpyxl

class ExcelWriter():
    def __init__(self, file_path):
        self.file = file_path
        self.work_book = openpyxl.load_workbook( self.file )
        self.work_sheet = self.work_book.worksheets[1]
        self.row_count = self.work_sheet.max_row
        self.col_count = self.work_sheet.max_column


    def update(self, data):
        print("data--called", self.row_count)

        for row_index in range( self.row_count ):

            val = self.work_sheet.cell(row=(row_index+1), column=2).value
            if val is None:
                continue
            for line in data:
                #print(f"---{row_index}---{val}--{line.excel_name}**{line.name}")
                if line.excel_name == val:
                    print(line.to_json())
                    value = self.work_sheet.cell( row=( row_index + 1), column=3 ).value

                    if value is not None:
                        self.work_sheet.cell( row=( row_index + 1), column=3 ).value = float(value) + self.format_value( line )
                    else:
                        self.work_sheet.cell( row=( row_index + 1), column=3 ).value = self.format_value( line )

        self.work_book.save(self.file)
        self.work_book.close


    def format_value(self, line):
        try:
            val = ''.join(line.val_1.split(',')).replace(')', '').replace('|', '')
            if val == "Nil" or val == 'NIL' or val == '_•' or val == '•-' or val == '._•' or val == '.-':
                return 0
            print("Value*******>", val, line.name)
            return round( ( float(val) / 100 ), 2 )
        except:
            print(f"ERROR:-->Name:{line.name}-->Val:{line.val_1}")
            return float(0)
        #return round( ( float(val) / 1_00_00_000 ), 2 )
       

#data = [ {'name': 'Total non current assets', 'val_1': '9,612.07', 'val_2': '9,347.40', 'excel_name': 'Non-Current Investments'}, {'name': 'Total non current assets', 'val_1': '9,612.07', 'val_2': '9,347.40', 'excel_name': 'Non-Current Investments'}] 
#file_path = "/Users/shravanc/flask/aditya_birla/nirma_sample.xlsx"

#ew = ExcelWriter(file_path)
#ew.update(data)
